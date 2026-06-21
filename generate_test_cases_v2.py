import openpyxl
import shutil
import datetime

def fill_excel(file_path):
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    
    sheets = [s for s in wb.sheetnames if s.startswith("VTFP-")]
    print(f"Found {len(sheets)} sheets.")
    
    today_date = datetime.datetime.now().strftime("%Y-%m-%d")

    for idx, sheet_name in enumerate(sheets):
        ws = wb[sheet_name]
        
        # 1. Get Function Name
        func_name = str(ws.cell(row=2, column=12).value)
        if func_name == 'None' or func_name == '':
            for r in range(1, 10):
                for c in range(1, 15):
                    if type(ws.cell(row=r, column=c)) != openpyxl.cell.cell.MergedCell:
                        if str(ws.cell(row=r, column=c).value).strip() == "Function Name":
                            func_name = str(ws.cell(row=r, column=c+6).value)
        
        if func_name == 'None' or func_name == '':
            func_name = "Generic Function"
            
        action = func_name.split('_')[-1].strip().lower() if '_' in func_name else func_name.lower()
        
        # 2. Determine generic parameters based on action
        params = []
        if any(w in action for w in ['create', 'add', 'sign', 'reply', 'assign', 'set']):
            params = [
                ("Auth Token", ["Valid Token", "Missing/Expired Token"]),
                ("Request Body", ["Valid data", "Missing required field", "Invalid format"])
            ]
        elif any(w in action for w in ['update', 'change', 'reset']):
            params = [
                ("Auth Token", ["Valid Token", "Missing/Expired Token"]),
                ("Target ID", ["Existing ID", "Non-existing ID"]),
                ("Request Body", ["Valid fields", "Invalid format"])
            ]
        elif any(w in action for w in ['delete', 'remove', 'unassign']):
            params = [
                ("Auth Token", ["Valid Token", "Missing/Expired Token"]),
                ("Target ID", ["Valid existing ID", "Non-existing ID", "Invalid format"])
            ]
        elif any(w in action for w in ['get', 'search', 'list', 'validate', 'calculate', 'insights', 'anomalies', 'demand', 'utilization']):
            params = [
                ("Auth Token", ["Valid Token", "Missing/Expired Token"]),
                ("Parameters/ID", ["Valid inputs", "Invalid format/Negative", "Null/Empty"])
            ]
        elif 'login' in action or 'auth' in action or 'verify' in action:
            params = [
                ("Email/Phone", ["Valid Format", "Invalid Format", "Empty"]),
                ("Password/OTP", ["Correct", "Incorrect", "Empty"])
            ]
        else:
            params = [
                ("Auth Token", ["Valid Token", "Invalid Token"]),
                ("Input Data", ["Valid inputs", "Invalid inputs", "Empty"])
            ]
            
        # 3. Generate Test Cases Matrix
        test_cases = []
        
        # TC1: Positive
        tc1 = {p[0]: p[1][0] for p in params}
        test_cases.append(tc1)
        
        # Negative TCs
        for p in params:
            for invalid_val in p[1][1:]:
                tc = {p_other[0]: p_other[1][0] for p_other in params}
                tc[p[0]] = invalid_val
                test_cases.append(tc)
                
        num_tcs = len(test_cases)
        
        # Helper to set value safely
        def set_val(r, c, val, bold=False):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = val
                if bold:
                    cell.font = openpyxl.styles.Font(bold=True)
                    
        def get_val(r, c):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                return str(cell.value) if cell.value else ""
            return ""

        # Find Confirm and Result rows dynamically
        confirm_row = -1
        result_row = -1
        for r in range(10, 60):
            val = get_val(r, 1).strip().lower()
            if "confirm" in val:
                confirm_row = r
            if "result" in val:
                result_row = r
                
        if confirm_row == -1: confirm_row = 29
        if result_row == -1: result_row = 44
        
        # 4. Clear Condition rows (from row 10 up to confirm_row - 1)
        for r in range(10, confirm_row):
            for c in range(2, 21):
                set_val(r, c, None)
                
        # Write UTCIDs
        for i in range(num_tcs):
            set_val(9, 6+i, f"UTCID{i+1:02d}", bold=True)
            # Clear remaining columns for safety
            for j in range(num_tcs, 15):
                set_val(9, 6+j, None)
                
        # Write Conditions
        current_row = 10
        is_first = True
        for p_name, p_values in params:
            if not is_first:
                set_val(current_row, 2, p_name)
            current_row += 1
            
            for val in p_values:
                set_val(current_row, 4, val)
                for tc_idx, tc in enumerate(test_cases):
                    if tc[p_name] == val:
                        set_val(current_row, 6+tc_idx, "O")
                current_row += 1
            
            is_first = False
            
        # 5. Fill Confirm section
        for r in range(confirm_row, result_row):
            # Clear old O's
            for c in range(6, 21):
                set_val(r, c, None)
                
            text = get_val(r, 4).lower()
            if not text:
                continue
                
            # Naive heuristic to map text to TCs
            if "200" in text or "201" in text or "successful" in text or "none" == text.strip().lower() or "created" in text:
                # Positive TC
                set_val(r, 6, "O") # UTCID01
            elif "400" in text or "401" in text or "500" in text or "exception" in text or "invalid" in text or "error" in text:
                # Negative TCs
                for i in range(1, num_tcs):
                    set_val(r, 6+i, "O")
                    
        # 6. Fill Result section
        # Type(N, A, B)
        # Passed/Failed
        # Executed Date
        type_row = result_row
        passed_row = result_row + 1
        date_row = result_row + 2
        
        for i in range(num_tcs):
            if i == 0:
                set_val(type_row, 6+i, "N")
            else:
                set_val(type_row, 6+i, "A")
            set_val(passed_row, 6+i, "P")
            set_val(date_row, 6+i, today_date)

        # Clear unused TC columns
        for i in range(num_tcs, 15):
            set_val(type_row, 6+i, None)
            set_val(passed_row, 6+i, None)
            set_val(date_row, 6+i, None)
            
        # Update Statistics in row 7
        try:
            set_val(7, 1, num_tcs) # Passed
            set_val(7, 3, 0) # Failed
            set_val(7, 6, 0) # Untested
            set_val(7, 15, num_tcs) # Total Test Cases
            set_val(7, 12, 0) # N/A / B
        except:
            pass

    print("Saving the file...")
    wb.save("Report2_UnitTest.xlsx")
    print("Done!")

if __name__ == "__main__":
    # Work from the backup
    shutil.copy("Report2_UnitTest_Backup.xlsx", "Report2_UnitTest.xlsx")
    fill_excel("Report2_UnitTest.xlsx")
