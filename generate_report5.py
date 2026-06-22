import openpyxl
from openpyxl.styles import Font
import re

def run():
    file_path = "Report5_TestReport.xlsx"
    wb = openpyxl.load_workbook(file_path)
    
    ws_stats = wb["Test Statistics"]
    
    modules = []
    for r in range(11, 41):
        mod_name = ws_stats.cell(row=r, column=3).value
        if mod_name:
            modules.append((r, str(mod_name)))
            
    categories = [
        ("Functional Positive", "Verify successful operation with valid inputs", "1. Navigate to module\n2. Enter valid data\n3. Submit", "Operation successful, data saved"),
        ("Functional Negative", "Verify operation fails with invalid inputs", "1. Navigate to module\n2. Enter invalid data\n3. Submit", "Error message displayed, operation rejected"),
        ("Boundary", "Verify operation at boundary limits", "1. Navigate to module\n2. Enter data at max/min limits\n3. Submit", "System handles boundary values correctly"),
        ("Edge Case", "Verify system behavior in uncommon scenarios", "1. Setup edge case condition\n2. Execute action", "System handles edge case gracefully without crashing"),
        ("Validation", "Verify mandatory fields and data types", "1. Leave mandatory fields blank\n2. Submit", "Validation errors for all mandatory fields"),
        ("Security", "Verify unauthorized access prevention", "1. Attempt action without proper permissions\n2. Execute", "Access denied, unauthorized error displayed")
    ]
    
    headers = [
        "Test Case ID", "Test Case Description", "Test Case Procedure", "Expected Result",
        "Pre-condition", "Round 1", "Test Date", "Tester", "Round 2", "Test Date", "Tester",
        "Round 3", "Test Date", "Tester", "Note"
    ]
    
    total_tcs = 0
    markdown_output = "# Báo cáo sinh Test Case - Report5_TestReport\n\n"
    markdown_output += "## 1. Danh sách sheet đã được tạo\n"
    
    for r, mod_name in modules:
        safe_sheet_name = re.sub(r'[\\\\/*?:\\[\\]]', '', mod_name)[:31].strip()
        markdown_output += f"- {safe_sheet_name}\n"
        
    markdown_output += "\n## 2. Danh sách Test Case của từng sheet\n"
        
    for r, mod_name in modules:
        safe_sheet_name = re.sub(r'[\\\\/*?:\\[\\]]', '', mod_name)[:31].strip()
        
        if safe_sheet_name in wb.sheetnames:
            del wb[safe_sheet_name]
            
        ws = wb.create_sheet(title=safe_sheet_name)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            
        mod_prefix = ''.join([w[0].upper() for w in re.sub(r'[^a-zA-Z ]', '', mod_name).split()])[:4]
        if len(mod_prefix) < 2:
            mod_prefix = re.sub(r'[^a-zA-Z]', '', mod_name).upper()[:4]
            
        tc_list = []
        markdown_output += f"\n### Module: {mod_name} (Sheet: {safe_sheet_name})\n"
        markdown_output += "| Test Case ID | Loại | Mô tả |\n|---|---|---|\n"
        
        for i, cat in enumerate(categories, 1):
            tc_id = f"TC-{mod_prefix}-{i:03d}"
            tc_desc = f"[{cat[0]}] {cat[1]} for {mod_name}"
            tc_proc = cat[2]
            tc_exp = cat[3]
            tc_pre = "System is accessible, user logged in if required"
            
            row_data = [
                tc_id, tc_desc, tc_proc, tc_exp, tc_pre,
                "Untested", "", "", "Untested", "", "", "Untested", "", "", ""
            ]
            
            for col, val in enumerate(row_data, 1):
                ws.cell(row=i+1, column=col).value = val
                
            tc_list.append(tc_id)
            total_tcs += 1
            
            markdown_output += f"| {tc_id} | {cat[0]} | {tc_desc} |\n"
            
        num_tcs = len(categories)
        ws_stats.cell(row=r, column=8).value = num_tcs
        
    markdown_output += f"\n## 3. Tổng số Test Case toàn hệ thống: {total_tcs}\n"
    
    wb.save(file_path)
    print("Excel saved successfully.")
    
    with open("test_cases_summary.md", "w", encoding="utf-8") as f:
        f.write(markdown_output)

if __name__ == "__main__":
    run()
