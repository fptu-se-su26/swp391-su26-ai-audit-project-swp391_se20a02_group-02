"""
mastp/tools/excel_exporter.py
Exports Test Cases to the standard LuxeWay TestReport xlsx format.
"""
from __future__ import annotations

from typing import List
from datetime import datetime


# ─── Column configuration ──────────────────────────────────────────────────────
COLUMNS = [
    "Test Case ID",
    "Test Case Description",
    "Test Case Procedure",
    "Expected Result",
    "Pre-condition",
    "Round 1",
    "Test Date",
    "Tester",
    "Round 2",
    "Test Date ",        # trailing space intentional — matches original template
    "Tester ",
    "Round 3",
    "Test Date  ",
    "Tester  ",
    "Note",
]

# ─── Color palette ─────────────────────────────────────────────────────────────
COLORS = {
    "header_fill":    "1F2937",   # dark slate
    "header_font":    "FFFFFF",
    "pass_fill":      "D1FAE5",   # green-100
    "fail_fill":      "FEE2E2",   # red-100
    "untested_fill":  "F3F4F6",   # gray-100
    "p0_fill":        "FEF3C7",   # amber-100  — P0 row highlight
    "p1_fill":        "EDE9FE",   # violet-100
    "alt_fill":       "F9FAFB",   # subtle alternating row
    "border_color":   "D1D5DB",
}

TYPE_COLORS = {
    "Functional Positive": "DCFCE7",
    "Functional Negative": "FEE2E2",
    "Boundary":            "DBEAFE",
    "Edge Case":           "FEF9C3",
    "Security":            "FFE4E6",
    "Validation":          "E0E7FF",
    "Performance":         "F3E8FF",
}


def write_test_cases_to_excel(
    test_cases: List[dict],
    output_path: str,
    module_code: str,
    module_name: str,
) -> str:
    """
    Write test cases to xlsx.
    Creates one sheet named after the module.
    Applies styling, auto-width columns, header freeze.

    Returns the output_path on success.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = module_code[:31]  # Excel sheet name max 31 chars

    # ── Helper styles ─────────────────────────────────────────────────────
    thin = Side(style="thin", color=COLORS["border_color"])
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def cell_fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Module title row ──────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"📋 {module_name} — Test Case Report"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = header_fill(COLORS["header_fill"])
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 28

    # ── Stats row ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
    stats_cell = ws["A2"]
    type_counts = {}
    for tc in test_cases:
        t = tc.get("test_case_type", "Other")
        type_counts[t] = type_counts.get(t, 0) + 1
    stats_cell.value = (
        f"Total: {len(test_cases)} TCs  |  "
        + "  |  ".join(f"{k}: {v}" for k, v in type_counts.items())
        + f"  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    stats_cell.font = Font(italic=True, size=10, color="374151")
    stats_cell.fill = cell_fill("F9FAFB")
    stats_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Column headers (row 3) ────────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=10, color=COLORS["header_font"])
        cell.fill = header_fill("374151")
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[3].height = 22

    # ── Freeze panes at row 4 ─────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Test case rows ────────────────────────────────────────────────────
    for row_offset, tc in enumerate(test_cases):
        row = 4 + row_offset
        tc_type = tc.get("test_case_type", "Functional Positive")
        row_color = TYPE_COLORS.get(tc_type, "FFFFFF")

        # Override for P0
        if tc.get("priority") == "P0":
            row_color = COLORS["p0_fill"]

        row_fill = cell_fill(row_color)

        values = [
            tc.get("test_case_id", ""),
            tc.get("test_case_description", ""),
            tc.get("test_case_procedure", ""),
            tc.get("expected_result", ""),
            tc.get("pre_condition", ""),
            tc.get("round1_result", "Untested"),
            tc.get("round1_date", ""),
            tc.get("round1_tester", ""),
            tc.get("round2_result", "Untested"),
            tc.get("round2_date", ""),
            tc.get("round2_tester", ""),
            tc.get("round3_result", "Untested"),
            tc.get("round3_date", ""),
            tc.get("round3_tester", ""),
            tc.get("note", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = wrap_align
            cell.border = thin_border
            cell.font = Font(size=9)

        ws.row_dimensions[row].height = max(
            60,
            min(len(tc.get("test_case_procedure", "")) // 3, 200)
        )

    # ── Column widths ─────────────────────────────────────────────────────
    col_widths = {
        1: 14,   # TC ID
        2: 42,   # Description
        3: 60,   # Procedure
        4: 45,   # Expected
        5: 40,   # Pre-condition
        6: 14,   # Round 1
        7: 14,   # Date
        8: 18,   # Tester
        9: 14,   # Round 2
        10: 14,
        11: 18,
        12: 14,  # Round 3
        13: 14,
        14: 18,
        15: 30,  # Note
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    return output_path


def write_multi_module_excel(
    test_cases_by_module: dict,          # {module_code: (module_name, [tc_dicts])}
    output_path: str,
) -> str:
    """
    Write multiple modules into separate sheets in one workbook.
    Also creates a 'Test Statistics' summary sheet.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Statistics sheet ─────────────────────────────────────────────────
    stats_ws = wb.create_sheet("Test Statistics", 0)
    stats_ws["A1"] = "Module Code"
    stats_ws["B1"] = "Module Name"
    stats_ws["C1"] = "Number of Test Cases"
    stats_ws["D1"] = "Pass"
    stats_ws["E1"] = "Fail"
    stats_ws["F1"] = "Untested"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col in "ABCDEF":
        c = stats_ws[f"{col}1"]
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    total_tcs = 0
    for stat_row, (module_code, (module_name, tcs)) in enumerate(
        test_cases_by_module.items(), start=2
    ):
        count = len(tcs)
        total_tcs += count
        stats_ws.cell(row=stat_row, column=1, value=module_code)
        stats_ws.cell(row=stat_row, column=2, value=module_name)
        stats_ws.cell(row=stat_row, column=3, value=count)
        stats_ws.cell(row=stat_row, column=4, value=0)
        stats_ws.cell(row=stat_row, column=5, value=0)
        stats_ws.cell(row=stat_row, column=6, value=count)

    # Total row
    total_row = 2 + len(test_cases_by_module)
    stats_ws.cell(row=total_row, column=2, value="TOTAL")
    stats_ws.cell(row=total_row, column=3, value=total_tcs)
    total_font = Font(bold=True)
    stats_ws.cell(row=total_row, column=2).font = total_font
    stats_ws.cell(row=total_row, column=3).font = total_font

    for col_letter, width in [("A", 14), ("B", 40), ("C", 22), ("D", 10), ("E", 10), ("F", 12)]:
        stats_ws.column_dimensions[col_letter].width = width

    # ── Per-module sheets ────────────────────────────────────────────────
    for module_code, (module_name, tcs) in test_cases_by_module.items():
        # Write to temp single-module workbook, then copy sheet
        from openpyxl import load_workbook
        import tempfile, os

        tmp = tempfile.mktemp(suffix=".xlsx")
        write_test_cases_to_excel(tcs, tmp, module_code, module_name)
        tmp_wb = load_workbook(tmp)
        src_ws = tmp_wb.active

        # Copy to main workbook
        new_ws = wb.create_sheet(module_code[:31])
        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.alignment = cell.alignment.copy()
                    new_cell.border = cell.border.copy()

        for col in src_ws.column_dimensions:
            new_ws.column_dimensions[col].width = src_ws.column_dimensions[col].width
        for row_num in src_ws.row_dimensions:
            new_ws.row_dimensions[row_num].height = src_ws.row_dimensions[row_num].height

        new_ws.freeze_panes = "A4"
        os.unlink(tmp)
        tmp_wb.close()

    wb.save(output_path)
    return output_path
