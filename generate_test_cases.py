import openpyxl
from copy import copy

def fill_excel(file_path):
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    
    # We will process all sheets starting with VTFP-
    sheets = [s for s in wb.sheetnames if s.startswith("VTFP-")]
    print(f"Found {len(sheets)} sheets.")
    
    for idx, sheet_name in enumerate(sheets):
        ws = wb[sheet_name]
        
        # 1. Get Function Name
        func_name = str(ws.cell(row=2, column=12).value)
        if func_name == 'None' or func_name == '':
            for r in range(1, 10):
                for c in range(1, 15):
                    if type(ws.cell(row=r, column=c)) != openpyxl.cell.cell.MergedCell:
                        if str(ws.cell(row=r, column=c).value).strip() == "Function Name":
                            func_name = str(ws.cell(row=r, column=c+6).value)
        
        if func_name == 'None' or func_name == '':
            func_name = "Generic Function"
            
        action = func_name.split('_')[-1].strip().lower() if '_' in func_name else func_name.lower()
        
        # 2. Determine generic parameters based on action
        params = []
        if any(w in action for w in ['create', 'add', 'sign', 'reply', 'assign', 'set']):
            params = [
                ("Auth Token", ["Valid Token", "Missing/Expired Token"]),
                ("Request Body", ["All fields valid", "Missing required field", "Invalid format"])
            ]
        elif any(w in action for w in ['update', 'change', 'reset']):
            params = [
                ("Auth Token", ["Valid Token", "Missing/Expired Token"]),
                ("Target ID", ["Existing ID", "Non-existing ID"]),
                ("Request Body", ["Valid fields", "Invalid data format"])
            ]
        elif any(w in action for w in ['delete', 'remove', 'unassign']):
            params = [
                ("Auth Token", ["Valid Token", "Missing/Expired Token"]),
                ("Target ID", ["Valid existing ID", "Non-existing ID", "Invalid format ID"])
            ]
        elif any(w in action for w in ['get', 'search', 'list', 'validate', 'calculate', 'insights', 'anomalies', 'demand', 'utilization']):
            params = [
                ("Auth Token", ["Valid Token", "Missing/Expired Token"]),
                ("Parameters/ID", ["Valid inputs", "Invalid format", "Null/Empty"])
            ]
        elif 'login' in action or 'auth' in action or 'verify' in action:
            params = [
                ("Email/Phone", ["Valid Format", "Invalid Format", "Empty"]),
                ("Password/OTP", ["Correct", "Incorrect", "Empty"])
            ]
        else: # default generic
            params = [
                ("Auth Token", ["Valid", "Invalid"]),
                ("Input Data", ["Valid", "Invalid", "Empty"])
            ]
            
        # 3. Generate Test Cases Matrix
        test_cases = []
        
        # TC1: All valid
        tc1 = {p[0]: p[1][0] for p in params}
        test_cases.append(tc1)
        
        # Negative TCs
        for p in params:
            for invalid_val in p[1][1:]:
                tc = {p_other[0]: p_other[1][0] for p_other in params}
                tc[p[0]] = invalid_val
                test_cases.append(tc)
                
        num_tcs = len(test_cases)
        
        # Helper to set value safely
        def set_val(r, c, val, bold=False):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = val
                if bold:
                    cell.font = openpyxl.styles.Font(bold=True)
        
        # 4. Clear old data from row 9 to 40, col 1 to 20
        for r in range(9, 41):
            for c in range(1, 21):
                set_val(r, c, None)
                
        # 5. Write the Matrix
        # Row 9: Headers
        for i in range(num_tcs):
            set_val(9, 6+i, f"UTCID{i+1:02d}", bold=True)
            
        current_row = 10
        set_val(current_row, 1, "Condition")
        set_val(current_row, 2, "Precondition")
        current_row += 1
        
        is_first = True
        for p_name, p_values in params:
            if not is_first:
                set_val(current_row, 2, p_name)
            current_row += 1
            
            for val in p_values:
                set_val(current_row, 4, val)
                for tc_idx, tc in enumerate(test_cases):
                    if tc[p_name] == val:
                        set_val(current_row, 6+tc_idx, "O")
                current_row += 1
            
            is_first = False
            
        # Update Statistics in row 7
        set_val(7, 1, 0) # Passed
        set_val(7, 3, 0) # Failed
        set_val(7, 6, num_tcs) # Untested
        set_val(7, 15, num_tcs) # Total Test Cases
        set_val(7, 12, 0) # N/A / B

    print("Saving the file...")
    wb.save(file_path)
    print("Done!")

if __name__ == "__main__":
    fill_excel("Report2_UnitTest.xlsx")
